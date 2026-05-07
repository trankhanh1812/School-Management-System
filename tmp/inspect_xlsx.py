from openpyxl import load_workbook

path = r'D:\School\School-Management-System\school-management\src\main\resources\templates\Thời khóa biểu.xlsx'
wb = load_workbook(path)
print(wb.sheetnames)
for ws in wb.worksheets:
    print('SHEET', ws.title, ws.max_row, ws.max_column)
    for row in ws.iter_rows():
        vals = [c.value for c in row]
        if any(v is not None for v in vals):
            print('ROW', row[0].row, vals)
    print('---')
