from openpyxl import load_workbook
from openpyxl.styles import Alignment

path = r'D:\School\School-Management-System\PGNV.xlsx'
wb = load_workbook(path)
ws = wb.worksheets[0]

section3 = {
    'A26': 'Phân tích nghiệp vụ quản lý trường THPT; mô hình hóa các thực thể User/Role/Student/Teacher/Class/Subject/Exam/Score/Attendance/Timetable/Notification; nắm quy trình RBAC, phân công giảng dạy, nhập điểm, duyệt điểm, điểm danh và thời khóa biểu.',
    'A30': 'Java 17, Spring Boot 3.5.x, Spring Data JPA, Spring Security, JWT, PostgreSQL 15+, Next.js 16, React 19, TypeScript, TailwindCSS 4, REST API, Maven.',
    'A34': 'Kỹ năng phân tích yêu cầu, thiết kế ERD/schema, xây dựng API/DTO/validation, tích hợp FE-BE, kiểm thử, debug và quản lý tiến độ theo module.',
    'A36': 'Một hệ thống quản lý học sinh THPT chạy trên web, hỗ trợ auth, user/role, student, teacher/class, subject/teaching assignment, exam/score, attendance, timetable và notification.',
    'A40': 'Số hóa và tập trung hóa nghiệp vụ quản lý trường học, giảm thao tác thủ công, hạn chế sai sót nhập liệu, tăng khả năng phân quyền và truy vết dữ liệu.',
}

section4 = {
    'H45': 1, 'J45': 1,
    'A47': 'Tìm hiểu tổng quan bài toán, khảo sát yêu cầu nghiệp vụ, xác định phạm vi hệ thống, phân tích chức năng chính và thiết kế sơ bộ CSDL/ERD.',
    'H50': 1, 'J50': 2,
    'A52': 'Tìm hiểu và thiết lập công nghệ nền tảng: Spring Boot, PostgreSQL, Next.js, TailwindCSS, cấu trúc project, kiểm tra kết nối FE-BE và chuẩn hóa môi trường phát triển.',
    'H55': 2, 'J55': 2,
    'A57': 'Xây dựng nền tảng xác thực và phân quyền: entity User/Role, repository/service, API login/register, JWT, Spring Security, RBAC, màn hình đăng nhập và kiểm thử luồng auth.',
    'H60': 3, 'J60': 7,
    'A62': 'Phát triển các module nghiệp vụ chính của hệ thống gồm Student, Teacher, Class, Subject, Teaching Assignment, Exam, Score, Attendance, Timetable và Notification; đồng thời hoàn thiện giao diện và tích hợp API.',
    'H65': 8, 'J65': 8,
    'A67': 'Kiểm thử end-to-end, sửa lỗi, tối ưu CSDL và truy vấn, cải thiện UI/UX, viết báo cáo, làm slide và chuẩn bị demo sản phẩm.',
}

for cell, value in {**section3, **section4}.items():
    ws[cell] = value
    ws[cell].alignment = Alignment(wrap_text=True, vertical='top')

# Preserve the existing template style while making the content readable.
for row in [26, 30, 34, 36, 40, 47, 52, 57, 62, 67]:
    ws.row_dimensions[row].height = 42

for row in [45, 50, 55, 60, 65]:
    ws.row_dimensions[row].height = 24

wb.save(path)
print('Updated', path)
